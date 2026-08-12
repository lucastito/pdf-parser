"""Vocabulário de campo esperado — para achar valor em texto sem tabela.

A fonte típica é a planilha de schema do destino (ex.: o input de um sistema
de um cenário corporativo). É a mesma categoria de configuração de negócio que já embasa
`esquema` e `mapeamento`: o núcleo não conhece nome de campo nenhum — só sabe
procurar o que foi declarado. Nenhum nome de domínio vive neste módulo; ele só
sabe **ler** um vocabulário, de onde quer que ele venha.

A planilha de schema não é o formato de saída — é só a lista do que procurar.
O destino real (produto final da extração) pode ser qualquer coisa: CSV, JSON,
ou o payload de uma chamada a outro sistema. Este módulo não presume nenhum
dos três.
"""

from __future__ import annotations

import re
import xml.etree.ElementTree as ET
import zipfile
from dataclasses import dataclass, field
from pathlib import Path

__all__ = ["CampoEsperado", "carregar_campos_do_xlsx"]

_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_X14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
_NS_XM = "http://schemas.microsoft.com/office/excel/2006/main"
_NS_REL = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

_CELULA = re.compile(r"^\$?([A-Z]+)\$?(\d+)$")
_REFERENCIA_DE_FAIXA = re.compile(
    r"^(?:'([^']+)'|([^!]+))!\$?[A-Z]+\$?(\d+):\$?[A-Z]+\$?(\d+)$"
)


@dataclass(frozen=True)
class CampoEsperado:
    """Um campo que o destino espera, e como reconhecê-lo em texto corrido."""

    nome: str
    sinonimos: tuple[str, ...] = field(default_factory=tuple)
    descricao: str | None = None
    unidade: str | None = None
    opcoes: tuple[str, ...] | None = None
    """As únicas respostas válidas, quando a planilha de schema declara uma
    lista suspensa para este campo. `None` quando o campo aceita valor livre
    ou quando nenhuma regra de validação foi encontrada — as duas situações
    são indistinguíveis a partir da planilha, e tratá-las como equivalentes é
    o comportamento conservador correto.

    **Limite medido, não hipotético:** a regra é casada por número de linha,
    não por coluna — `sqref` diz "estas linhas", e a validação pode ter sido
    declarada para uma coluna administrativa da planilha (ex.: uma marcação
    "obrigatório? Sim/Não") que não é a própria célula de valor do campo.
    Contra uma planilha real de um cenário corporativo isso produziu casos claramente errados
    (campo de texto livre recebendo `('Yes', 'No')`). Trate como candidato a
    confirmar, não como verdade — o mesmo espírito do achado sobre a coluna
    de unidade em planilhas grandes."""

    faixa: tuple[float, float] | None = None
    """Mínimo e máximo, quando a planilha declara uma faixa numérica válida.
    Só populado quando os dois limites são conhecidos — uma regra de "maior
    que X", sem teto, não teria como que valor citar como faixa, e fica de
    fora em vez de inventar um limite superior."""

    def rotulos(self) -> tuple[str, ...]:
        """Todos os textos que contam como menção a este campo."""
        return (self.nome, *self.sinonimos)


def carregar_campos_do_xlsx(
    caminho: str | Path,
    *,
    aba: str | None = None,
    abas: list[str] | None = None,
    coluna_nome: str = "PARAMETER",
    coluna_unidade: str = "UNIT",
) -> list[CampoEsperado]:
    """Lê uma lista de campos esperados de uma planilha de schema.

    A convenção assumida — uma linha de cabeçalho com nomes de coluna, e
    seções intercaladas que só preenchem a primeira célula da linha — é a de
    uma planilha de input de projeto genérica, não de nenhum produto
    específico: qualquer schema no mesmo formato serve. Uma linha de seção
    (`"RESERVOIR: FLUID PROPERTIES"`, célula A sozinha) é distinguida de uma
    linha de campo porque a coluna do nome do campo vem vazia nela — não por
    posição ou por padrão de texto.

    Quando a coluna de unidade existir no cabeçalho, o valor bruto entra em
    `CampoEsperado.unidade` sem validação — planilhas grandes o bastante têm
    seções cuja coluna de unidade traz outra coisa (coordenada, rótulo de
    duas linhas); declarar aqui uma unidade que não faz sentido é melhor que
    inventar uma heurística frágil para filtrá-la, e o limite fica registrado
    em vez de escondido.

    Além do nome, tenta recuperar **opções de lista suspensa** e **faixa
    numérica**, quando a planilha as declarar como validação de dado. O
    `openpyxl` não lê o formato de extensão (`x14:dataValidation`) que o
    Excel usa para a maioria das listas suspensas em planilhas grandes — a
    leitura aqui abre o `.xlsx` como o zip que ele é e lê o XML da aba
    diretamente. É enriquecimento **best-effort**: falhar a ler validação não
    impede a leitura dos campos em si.

    Args:
        aba: nome de uma única aba a ler. Mutuamente exclusivo com `abas`.
        abas: nomes de várias abas, mescladas na ordem dada — campo duplicado
            (mesmo nome) entra uma vez, na primeira aba em que aparece. Uma
            aba sem cabeçalho reconhecível, ou sem campo algum, é **ignorada
            com aviso** em vez de interromper as demais: planilhas grandes o
            bastante têm seções cuja estrutura não segue a convenção — e
            perder essa aba não deveria custar todas as outras.
        coluna_nome: cabeçalho da coluna com o nome do campo.
        coluna_unidade: cabeçalho da coluna com a unidade do campo, se houver.

    Levanta:
        ValueError: nem `aba` nem `abas` informado, os dois informados juntos,
            ou (no caso de uma única `aba`) ela não tem cabeçalho reconhecível
            nem campo algum. Com `abas`, o mesmo caso por aba é degradado a
            aviso — só levanta se **nenhuma** aba produzir campo.
    """
    if aba and abas:
        raise ValueError("informe 'aba' ou 'abas', não as duas")
    if not aba and not abas:
        raise ValueError("informe 'aba' (uma) ou 'abas' (várias)")

    import openpyxl

    workbook = openpyxl.load_workbook(str(caminho), data_only=True)

    if aba:
        return _carregar_de_uma_aba(workbook, caminho, aba, coluna_nome, coluna_unidade)

    campos: list[CampoEsperado] = []
    vistos: set[str] = set()
    ignoradas: list[str] = []
    for nome_aba in abas:
        try:
            novos = _carregar_de_uma_aba(
                workbook, caminho, nome_aba, coluna_nome, coluna_unidade
            )
        except ValueError as erro:
            ignoradas.append(nome_aba)
            print(f"  vocabulário: aba {nome_aba!r} ignorada — {erro}")
            continue
        for campo in novos:
            if campo.nome in vistos:
                continue
            vistos.add(campo.nome)
            campos.append(campo)

    if not campos:
        raise ValueError(
            f"nenhuma das {len(abas)} aba(s) produziu campo algum "
            f"({len(ignoradas)} ignorada(s): {', '.join(ignoradas)})"
        )

    return campos


def _carregar_de_uma_aba(
    workbook, caminho: str | Path, aba: str, coluna_nome: str, coluna_unidade: str
) -> list[CampoEsperado]:
    if aba not in workbook.sheetnames:
        raise ValueError(f"aba {aba!r} não existe nesta planilha")

    planilha = workbook[aba]
    linhas = list(planilha.iter_rows(values_only=True))

    indice_cabecalho = None
    cabecalho: list[str] = []
    for indice, linha in enumerate(linhas):
        celulas = [str(c).strip() if c is not None else "" for c in linha]
        if coluna_nome in celulas:
            cabecalho = celulas
            indice_cabecalho = indice
            break

    if indice_cabecalho is None:
        raise ValueError(
            f"aba {aba!r} não tem uma linha de cabeçalho com a coluna {coluna_nome!r}"
        )

    indice_nome = cabecalho.index(coluna_nome)
    indice_descricao = cabecalho.index("DESCRIPTION") if "DESCRIPTION" in cabecalho else None
    indice_unidade = cabecalho.index(coluna_unidade) if coluna_unidade in cabecalho else None

    try:
        validacoes = _ler_validacoes_da_aba(caminho, workbook, aba)
    except Exception:  # noqa: BLE001 — enriquecimento best-effort, nunca bloqueia
        validacoes = {}

    campos: list[CampoEsperado] = []
    vistos: set[str] = set()
    for deslocamento, linha in enumerate(linhas[indice_cabecalho + 1 :]):
        numero_no_excel = indice_cabecalho + 2 + deslocamento  # 1-indexado, após o cabeçalho

        nome = linha[indice_nome] if indice_nome < len(linha) else None
        if not nome or not str(nome).strip():
            continue  # linha de seção, ou célula vazia — não é campo

        nome = str(nome).strip()
        if nome in vistos:
            continue
        vistos.add(nome)

        descricao = None
        if indice_descricao is not None and indice_descricao < len(linha):
            bruta = linha[indice_descricao]
            descricao = str(bruta).strip() if bruta else None

        unidade = None
        if indice_unidade is not None and indice_unidade < len(linha):
            bruta = linha[indice_unidade]
            unidade = str(bruta).strip() if bruta else None

        regra = validacoes.get(numero_no_excel, {})
        campos.append(
            CampoEsperado(
                nome=nome,
                descricao=descricao,
                unidade=unidade,
                opcoes=regra.get("opcoes"),
                faixa=regra.get("faixa"),
            )
        )

    if not campos:
        raise ValueError(f"aba {aba!r} não tem campo algum depois do cabeçalho")

    return campos


def _mapear_planilha_para_arquivo(caminho: str | Path) -> dict[str, str]:
    """Nome da aba → caminho do XML dentro do `.xlsx`.

    A ordem das abas no arquivo não garante bater com a numeração dos
    arquivos `sheetN.xml` — o mapeamento real vive em `xl/workbook.xml`
    (nome → r:id) e `xl/_rels/workbook.xml.rels` (r:id → arquivo).
    """
    with zipfile.ZipFile(caminho) as z:
        workbook_xml = z.read("xl/workbook.xml")
        rels_xml = z.read("xl/_rels/workbook.xml.rels")

    raiz_wb = ET.fromstring(workbook_xml)
    rid_por_nome: dict[str, str] = {}
    sheets = raiz_wb.find(f"{{{_NS_MAIN}}}sheets")
    for sheet in sheets if sheets is not None else []:
        nome = sheet.get("name")
        rid = sheet.get(f"{{{_NS_REL}}}id")
        if nome and rid:
            rid_por_nome[nome] = rid

    raiz_rel = ET.fromstring(rels_xml)
    arquivo_por_rid = {rel.get("Id"): rel.get("Target") for rel in raiz_rel if rel.get("Id")}

    return {
        nome: f"xl/{arquivo_por_rid[rid]}"
        for nome, rid in rid_por_nome.items()
        if rid in arquivo_por_rid
    }


def _ler_validacoes_da_aba(caminho: str | Path, workbook, aba: str) -> dict[int, dict]:
    """Regras de validação de dado desta aba, por linha (numeração do Excel).

    Cobre dois formatos: a validação "clássica" (`<dataValidation>`, que
    `openpyxl` também lê) e a de extensão (`<x14:dataValidation>`), que
    `openpyxl` ignora — e que é onde a maioria das listas suspensas mora em
    planilhas grandes. Sem ler o XML bruto, essas listas são invisíveis.
    """
    mapa = _mapear_planilha_para_arquivo(caminho)
    arquivo = mapa.get(aba)
    if arquivo is None:
        return {}

    with zipfile.ZipFile(caminho) as z:
        try:
            xml_bruto = z.read(arquivo)
        except KeyError:
            return {}

    raiz = ET.fromstring(xml_bruto)
    regras: dict[int, dict] = {}

    for validacao in raiz.iter(f"{{{_NS_MAIN}}}dataValidation"):
        _aplicar_validacao_classica(validacao, regras)

    for validacao in raiz.iter(f"{{{_NS_X14}}}dataValidation"):
        _aplicar_validacao_de_extensao(validacao, workbook, regras)

    return regras


def _aplicar_validacao_classica(elemento, regras: dict[int, dict]) -> None:
    tipo = elemento.get("type")
    linhas = _linhas_do_sqref(elemento.get("sqref", ""))
    if not linhas:
        return

    formula1 = _texto_do_filho(elemento, "formula1")
    formula2 = _texto_do_filho(elemento, "formula2")

    if tipo == "list" and formula1 and formula1.startswith('"') and formula1.endswith('"'):
        opcoes = tuple(v.strip() for v in formula1[1:-1].split(",") if v.strip())
        if opcoes:
            for linha in linhas:
                regras[linha] = {"opcoes": opcoes}
    elif tipo in ("whole", "decimal") and elemento.get("operator", "between") == "between":
        minimo, maximo = _para_numero(formula1), _para_numero(formula2)
        if minimo is not None and maximo is not None:
            for linha in linhas:
                regras[linha] = {"faixa": (minimo, maximo)}


def _aplicar_validacao_de_extensao(elemento, workbook, regras: dict[int, dict]) -> None:
    if elemento.get("type") != "list":
        return

    sqref = elemento.find(f"{{{_NS_XM}}}sqref")
    if sqref is None or not sqref.text:
        return
    linhas = _linhas_do_sqref(sqref.text)
    if not linhas:
        return

    formula1 = elemento.find(f"{{{_NS_X14}}}formula1")
    if formula1 is None:
        return
    referencia = formula1.find(f"{{{_NS_XM}}}f")
    if referencia is None or not referencia.text:
        return

    opcoes = _resolver_referencia_de_lista(workbook, referencia.text)
    if not opcoes:
        return
    for linha in linhas:
        regras[linha] = {"opcoes": opcoes}


def _resolver_referencia_de_lista(workbook, referencia: str) -> tuple[str, ...]:
    """Lê `"Aba!$B$57:$B$58"` como os valores de fato naquele intervalo."""
    casado = _REFERENCIA_DE_FAIXA.match(referencia.strip())
    if not casado:
        return ()

    aba = casado.group(1) or casado.group(2)
    linha_inicio, linha_fim = int(casado.group(3)), int(casado.group(4))
    coluna = re.search(r"\$?([A-Z]+)\$?\d+:", referencia)
    if aba not in workbook.sheetnames or not coluna:
        return ()

    planilha = workbook[aba]
    valores = []
    for linha in range(linha_inicio, linha_fim + 1):
        valor = planilha[f"{coluna.group(1)}{linha}"].value
        if valor is not None and str(valor).strip():
            valores.append(str(valor).strip())
    return tuple(valores)


def _linhas_do_sqref(sqref: str) -> list[int]:
    """As linhas (numeração do Excel) que um `sqref` cobre.

    `sqref` pode trazer vários intervalos separados por espaço
    (`"D5:D8 E38"`) — cada um vira as linhas que ele cobre, ignorando a
    coluna: o que importa aqui é casar com a linha do campo, não a célula
    exata onde o valor seria digitado.
    """
    linhas: set[int] = set()
    for token in sqref.split():
        if ":" in token:
            inicio, fim = token.split(":", 1)
            m1, m2 = _CELULA.match(inicio), _CELULA.match(fim)
            if m1 and m2:
                linhas.update(range(int(m1.group(2)), int(m2.group(2)) + 1))
        else:
            m = _CELULA.match(token)
            if m:
                linhas.add(int(m.group(2)))
    return sorted(linhas)


def _texto_do_filho(elemento, tag: str) -> str | None:
    filho = elemento.find(f"{{{_NS_MAIN}}}{tag}")
    return filho.text if filho is not None else None


def _para_numero(texto: str | None) -> float | None:
    if texto is None:
        return None
    try:
        return float(texto)
    except ValueError:
        return None
